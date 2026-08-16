# -*- coding: utf-8 -*-
import csv, os
import matcher as M
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

OUT = '/mnt/user-data/outputs'
SRC = 'user_sents.csv'

COLS = ['sentence_id','company_id','company_name','report_year','section','sentence_text',
        'matched_surface_forms','matched_core_ids','matched_pattern_ids','candidate_labels',
        'dedup_key','specific_technology_labels','product_labels','ai_general_anchor',
        'company_attribution','status_code','max_embedding_similarity','rule_confidence',
        'final_confidence','is_candidate','evidence_terms','notes','source_file','pdf_page',
        'report_page','subheading','text_type','original_sentence_id']

src = list(csv.DictReader(open(SRC, encoding='utf-8-sig')))
records = []
seq = 0
for r in src:
    res = M.match_sentence(r['句子'])
    if not res:
        continue
    seq += 1
    top = res.pop('_top_label')
    rec = {
        'sentence_id': f'CAND-{seq:05d}',
        'company_id': r['股票代码'],
        'company_name': r['公司'],
        'report_year': '2025',
        'section': r['章节'],
        'sentence_text': r['句子'],
        'dedup_key': f"{r['sentence_id']}+{top}",
        'source_file': r['源文件'],
        'pdf_page': r['PDF页序'],
        'report_page': r['页码'],
        'subheading': r['小标题'],
        'text_type': r['文本类型'],
        'original_sentence_id': r['sentence_id'],
    }
    rec.update(res)
    records.append(rec)

# ---- CSV ----
csv_path = os.path.join(OUT, 'candidate_sentences_v2.csv')
with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.DictWriter(f, fieldnames=COLS)
    w.writeheader()
    for rec in records:
        w.writerow({c: rec.get(c, '') for c in COLS})

# ---- XLSX ----
wb = Workbook()
ws = wb.active; ws.title = '候选句输出'
hf = Font(name='Arial', bold=True, color='FFFFFF'); fill = PatternFill('solid', fgColor='2F5597')
ws.append(COLS)
for c in ws[1]:
    c.font = hf; c.fill = fill; c.alignment = Alignment(vertical='center')
for rec in records:
    ws.append([rec.get(c, '') for c in COLS])
ws.freeze_panes = 'A2'
for col, wdt in {'F':60,'G':28,'H':20,'I':22,'J':26,'L':22,'M':20,'U':28,'V':22,'W':40,'Z':26}.items():
    ws.column_dimensions[col].width = wdt
for col in ['A','B','C','D','E']:
    ws.column_dimensions[col].width = 14

from collections import Counter
ws2 = wb.create_sheet('汇总')
def block(title, counter):
    ws2.append([title]); ws2[ws2.max_row][0].font = Font(name='Arial', bold=True)
    for k, v in counter.most_common():
        ws2.append([k, v])
    ws2.append([])
block('每章节候选数', Counter(r['section'] for r in records))
block('按文本类型', Counter(r['text_type'] for r in records))
lab = Counter()
for r in records:
    for l in r['candidate_labels'].split('; '):
        if l: lab[l]+=1
block('候选标签分布', lab)
block('状态分布', Counter(r['status_code'] or '(空)' for r in records))
block('归属分布', Counter(r['company_attribution'] for r in records))
for c in ('A','B'):
    ws2.column_dimensions[c].width = 26

xlsx_path = os.path.join(OUT, 'candidate_sentences_v2.xlsx')
wb.save(xlsx_path)

print('scanned:', len(src), '| candidates:', len(records))
print('by text_type:', dict(Counter(r['text_type'] for r in records)))
print('by section:', dict(Counter(r['section'] for r in records)))
print('CSV  ->', csv_path)
print('XLSX ->', xlsx_path)
